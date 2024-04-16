import os
import cv2
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import pandas as pd
from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def wound_segmentation(image_file):
    # Load the image
    image = cv2.imread(image_file)

    # Resize the image to reduce computation time
    image = cv2.resize(image, (400, 400))

    # Convert the image to the L*a*b* color space
    lab = cv2.cvtColor(image, cv2.COLOR_BGR2Lab)

    # Flatten the L*a*b* image
    lab_flat = lab.reshape((-1, 3))

    # Apply k-means clustering with 2 clusters (background and wound)
    kmeans = KMeans(n_clusters=2, random_state=42)
    kmeans.fit(lab_flat)

    # Assign each pixel to its corresponding cluster
    labels = kmeans.labels_.reshape(lab.shape[:2])

    # Determine the cluster with the highest average L* value (brightness)
    avg_l_values = [np.mean(lab[labels == i, 0]) for i in range(2)]
    background_label = np.argmax(avg_l_values)

    # Create a binary mask where the wound pixels are set to 255
    mask = np.where(labels == background_label, 0, 255).astype(np.uint8)

    # Perform morphological closing to fill small holes in the wound region
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)

    # Find contours in the mask
    contours, hierarchy = cv2.findContours(
        mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # Create a mask image for the wound region
    mask = np.zeros_like(mask)

    # Find the contour with the largest area (the wound region)
    max_contour = max(contours, key=cv2.contourArea)

    # Draw the contour on the mask image
    cv2.drawContours(mask, [max_contour], 0, 255, -1)
    # Create a binary mask where the wound pixels are set to 255
    mask = np.where(labels == background_label, 0, 255).astype(np.uint8)

    # Apply the mask to the original image
    masked_image = cv2.bitwise_and(image, image, mask=mask)

    # Display the original and masked images
    fig, axs = plt.subplots(1, 2, figsize=(10, 5))
    axs[0].imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    axs[0].set_title("Original")
    axs[1].imshow(cv2.cvtColor(masked_image, cv2.COLOR_BGR2RGB))
    axs[1].set_title("Masked")
    plt.show()

    # Return the mask and masked image
    return mask, masked_image


# Define image folder path
image_folder_path = "C:\\Users\\donne\\OneDrive\\Desktop\\wound-data-1"

# Create a list of image file paths
image_files = [os.path.join(image_folder_path, f) for f in os.listdir(
    image_folder_path) if f.endswith('.jpg') or f.endswith('.jpg')]


def rgb_to_hex(rgb):
    return '#{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])


# Define the color resolution for grouping
color_resolution = 10

# Loop over each image
for image_file in image_files:
    mask, masked_image = wound_segmentation(image_file)

    # Extract color data from the masked image
    colors = masked_image[mask > 0].reshape(-1, 3)

    # Quantize the color space
    quantized_colors = (colors // color_resolution) * color_resolution

    # Count the frequency of each quantized color
    color_counts = Counter(tuple(c) for c in quantized_colors)

    # Convert the quantized colors to hex ranges and store them along with their counts
    hex_ranges = []
    counts = []
    for color, count in color_counts.items():
        lower_bound = rgb_to_hex(color)
        upper_bound = rgb_to_hex(
            [min(c + color_resolution - 1, 255) for c in color])
        hex_range = f"{lower_bound}-{upper_bound}"
        hex_ranges.append(hex_range)
        counts.append(count)

    # Create a DataFrame from the hex color ranges and counts
    df = pd.DataFrame({"Hex Range": hex_ranges, "Count": counts})

    # Export the DataFrame to an Excel file
    excel_file = os.path.splitext(image_file)[0] + "_grouped_colors.xlsx"
    df.to_excel(excel_file, index=False, engine="openpyxl")

 # Create a DataFrame from the hex color ranges and counts
    df = pd.DataFrame({"Hex Range": hex_ranges, "Count": counts})

    # Filter the DataFrame to include only rows with counts over 100
    df = df[df["Count"] > 100]

    # Export the DataFrame to an Excel file
    excel_file = os.path.splitext(image_file)[0] + "_grouped_colors.xlsx"
    df.to_excel(excel_file, index=False, engine="openpyxl")

    # Filter the DataFrame to include only rows with counts over 100
    df = df[df["Count"] > 100]

# Create a pie chart of the top 5 color ranges, with the color shown in the chart being the upper and lower limit of the range
    top_5 = df.nlargest(5, "Count")
    upper_colors = [c.split("-")[1] for c in top_5["Hex Range"]]
    lower_colors = [c.split("-")[0] for c in top_5["Hex Range"]]
    labels = top_5["Hex Range"].tolist()

    # Show the pie chart with top 5 color ranges
    fig, axs = plt.subplots(2, 2, figsize=(10, 10))
    axs[0, 0].pie(top_5["Count"], labels=labels, colors=upper_colors)
    axs[0, 0].set_title("Top 5 Color Ranges (Upper Limit)")

    # Mask the image to only show the top 5 color ranges (using the upper limit)
    image = cv2.imread(image_file)
    upper_mask = np.zeros(image.shape[:2], np.uint8)
    for c_range in top_5["Hex Range"]:
        c1, c2 = c_range.split("-")
        c1_rgb = tuple(int(c1.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        c2_rgb = tuple(int(c2.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        c_mask = cv2.inRange(image, c1_rgb, c2_rgb)
        upper_mask |= c_mask

    # Apply the mask to the image (using the upper limit)
    upper_masked_image = cv2.bitwise_and(image, image, mask=upper_mask)

    # Display the masked image (using the upper limit)
    axs[1, 0].imshow(cv2.cvtColor(upper_masked_image, cv2.COLOR_BGR2RGB))
    axs[1, 0].set_title("Masked Image (Upper Limit)")

    # Show the pie chart with top 3 color ranges
    top_3 = df.nlargest(3, "Count")
    upper_colors_3 = [c.split("-")[1] for c in top_3["Hex Range"]]
    lower_colors_3 = [c.split("-")[0] for c in top_3["Hex Range"]]
    labels_3 = top_3["Hex Range"].tolist()

    axs[0, 1].pie(top_3["Count"], labels=labels_3, colors=upper_colors_3)
    axs[0, 1].set_title("Top 3 Color Ranges (Upper Limit)")

    # Mask the image to only show the top 5 color ranges (using the upper limit)
    upper_mask_3 = np.zeros(image.shape[:2], np.uint8)
    for c_range in top_3["Hex Range"]:
        c1, c2 = c_range.split("-")
        c1_rgb = tuple(int(c1.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        c2_rgb = tuple(int(c2.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        c_mask = cv2.inRange(image, c1_rgb, c2_rgb)
        upper_mask_3 |= c_mask

# Apply the mask to the image (using the upper limit)
    upper_masked_image_3 = cv2.bitwise_and(image, image, mask=upper_mask_3)

    # Display the masked image (using the lower limit)
    axs[1, 1].imshow(cv2.cvtColor(upper_masked_image_3, cv2.COLOR_BGR2RGB))
    axs[1, 1].set_title("Masked Image (Upper Limit)")

    plt.show()

    # Open the created Excel file using openpyxl
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Apply cell background colors according to the hex range
    for row in range(2, ws.max_row + 1):
        hex_range = ws.cell(row=row, column=1).value.split("-")
        lower_bound_fill = PatternFill(start_color=hex_range[0].lstrip(
            "#"), end_color=hex_range[0].lstrip("#"), fill_type="solid")
        upper_bound_fill = PatternFill(start_color=hex_range[1].lstrip(
            "#"), end_color=hex_range[1].lstrip("#"), fill_type="solid")
        ws.cell(row=row, column=3).fill = lower_bound_fill
        ws.cell(row=row, column=4).fill = upper_bound_fill

    # Save the modified workbook
    wb.save(excel_file)